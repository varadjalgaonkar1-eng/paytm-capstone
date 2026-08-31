import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

ledger_df = pd.read_csv("ledger.csv")
merchants_df = pd.read_csv("merchants.csv")

wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# 1. Merchants Sheet
ws_merchants = wb.create_sheet(title="Merchants")
for r in dataframe_to_rows(merchants_df, index=False, header=True):
    ws_merchants.append(r)

# 2. Fee Reference Table (HLOOKUP target)
ws_fees = wb.create_sheet(title="Fee_Tiers")
ws_fees.append(["Payment Method", "UPI", "Wallet", "Card", "Netbanking"])
ws_fees.append(["MDR Rate", 0.00, 0.015, 0.02, 0.012])

# 3. Transactions View Sheet
ws_txns = wb.create_sheet(title="Transactions_View")
ws_txns.append([
    "transaction_id",
    "user_id",
    "merchant_id",
    "transaction_time",
    "amount_inr",
    "payment_method",
    "status",
    "risk_score",
    "merchant_name",
    "category",
    "region",
    "mdr_rate",
    "high_value_merchant_day",
])

for i, row in ledger_df.iterrows():
    excel_row = i + 2
    ws_txns.append([
        row["transaction_id"],
        row["user_id"],
        row["merchant_id"],
        str(row["transaction_time"]),
        row["amount_inr"],
        row["payment_method"],
        row["status"],
        row["risk_score"],
        f'=IFERROR(VLOOKUP(C{excel_row}, Merchants!$A$2:$D$41, 2, FALSE), "Merchant not found")',
        f'=IFERROR(VLOOKUP(C{excel_row}, Merchants!$A$2:$D$41, 3, FALSE), "Merchant not found")',
        f'=IFERROR(VLOOKUP(C{excel_row}, Merchants!$A$2:$D$41, 4, FALSE), "Merchant not found")',
        f"=HLOOKUP(F{excel_row}, Fee_Tiers!$B$1:$E$2, 2, FALSE)",
        f'=IF(AND(E{excel_row}>5000, K{excel_row}<>">East"), "High-Value Merchant Day", "Standard")',
    ])

# 4. Summary & Pivot View Sheet
ws_summary = wb.create_sheet(title="Pivot_Summary")
pivot = (
    ledger_df.groupby(["merchant_id", "status"])
    .agg(total_amount_inr=("amount_inr", "sum"), txn_count=("amount_inr", "count"))
    .reset_index()
)

ws_summary.append(
    ["--- Pivot: Total Amount and Count by Merchant ID & Status ---"]
)
for r in dataframe_to_rows(pivot, index=False, header=True):
    ws_summary.append(r)

ws_summary.append([])
ws_summary.append(["--- Count vs Count-Unique (Active Transacted Days) ---"])
ws_summary.append(
    ["merchant_id", "total_transactions", "unique_transacted_days"]
)

ledger_df["txn_date"] = pd.to_datetime(ledger_df["transaction_time"]).dt.date
unique_days = (
    ledger_df.groupby("merchant_id")
    .agg(
        total_txns=("transaction_id", "count"),
        unique_days=("txn_date", "nunique"),
    )
    .head(10)
    .reset_index()
)

for r in dataframe_to_rows(unique_days, index=False, header=False):
    ws_summary.append(r)

wb.save("merchant_workbook.xlsx")
print("merchant_workbook.xlsx generated successfully.")
